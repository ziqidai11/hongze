#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from copy import copy

# —— 在这里填写你要操作的文件路径 ——
FILE_PATHS = [
    "宏观经济/eta/1.宏观经济_数据上传.xlsx",
    "wti模型3.0/eta/1.WTI_数据上传.xlsx",
    "汽柴煤油2.0/eta/1.汽柴煤油_数据上传.xlsx", 
    "汽柴煤油2.0/eta/2.汽柴煤油_数据上传.xlsx",
    "天然气/eta/1.天然气_数据上传.xlsx",
    "黑色/玻璃/eta/1.玻璃纯碱_数据上传.xlsx",
    "黑色/铁矿/eta/1.铁矿_数据上传.xlsx",
    "化工/eta/1.化工_数据上传.xlsx",
    "焦煤/eta/1.焦煤_数据上传.xlsx",
    "焦炭/eta/1.焦炭_数据上传.xlsx",
    "铝/eta/1.铝_数据上传.xlsx",
    "铜/eta/1.铜_数据上传.xlsx",
    "燃料油/eta/1.燃料油_数据上传.xlsx",
    "动力煤/eta/1.动力煤_数据上传.xlsx",
    "螺纹/eta/1.螺纹_数据上传.xlsx",
    "聚丙烯(PP)/eta/1.聚丙烯_数据上传.xlsx",
    "沥青/eta/1.沥青_数据上传.xlsx"
]

SUMMARY_SHEET = '列表页'
DETAIL_SHEET = '详情页'
DAILY_DATA_SHEET = '日度数据表'

# 新建目标工作簿
merged_wb = Workbook()

# 初始化三个sheet
merged_summary_ws = merged_wb.active
merged_summary_ws.title = SUMMARY_SHEET
merged_detail_ws = merged_wb.create_sheet(DETAIL_SHEET)
merged_daily_ws = merged_wb.create_sheet(DAILY_DATA_SHEET)

# 从第一个文件复制第一行固定话到合并结果中
if FILE_PATHS:
    first_wb = load_workbook(FILE_PATHS[0], data_only=False)
    first_ws = first_wb[SUMMARY_SHEET]
    first_row_max_col = first_ws.max_column
    
    # 复制第一行到合并结果的第一行
    for col in range(1, first_row_max_col + 1):
        source_cell = first_ws.cell(row=1, column=col)
        target_cell = merged_summary_ws.cell(row=1, column=col, value=source_cell.value)
        
        # 复制样式
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
    
    first_wb.close()

detail_col_offset = 1
daily_col_offset = 1

for i, file_path in enumerate(FILE_PATHS):
    print(f"[处理] {file_path}")
    wb = load_workbook(file_path, data_only=False)

    # ---- 列表页 ----
    ws_summary = wb[SUMMARY_SHEET]
    max_row_summary = ws_summary.max_row
    max_col_summary = ws_summary.max_column

    start_row = 2 if i == 0 else 3  # 第一个文件从第2行开始，其他文件从第3行开始（跳过第1行固定话和第2行表头）
    target_row = merged_summary_ws.max_row + 1

    for row in range(start_row, max_row_summary + 1):
        for col in range(1, max_col_summary + 1):
            source_cell = ws_summary.cell(row=row, column=col)
            target_cell = merged_summary_ws.cell(row=target_row, column=col, value=source_cell.value)

            # 复制样式
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format

        target_row += 1

    # ---- 详情页 ----
    ws_detail = wb[DETAIL_SHEET]
    max_row_detail = ws_detail.max_row
    max_col_detail = ws_detail.max_column

    for col_start in range(1, max_col_detail + 1, 5):
        # 检查这5列是否有数据
        has_data = False
        for row in range(1, max_row_detail + 1):
            for col in range(col_start, min(col_start + 5, max_col_detail + 1)):
                if ws_detail.cell(row=row, column=col).value is not None:
                    has_data = True
                    break
            if has_data:
                break
        
        # 只有当这5列有数据时才合并
        if has_data:
            for row in range(1, max_row_detail + 1):
                for col in range(col_start, min(col_start + 5, max_col_detail + 1)):
                    cell_value = ws_detail.cell(row=row, column=col).value
                    target_cell = merged_detail_ws.cell(row=row, column=detail_col_offset + (col - col_start), value=cell_value)
                    # 设置第5列为百分比格式
                    if (col - col_start) == 4:  # 第5列
                        target_cell.number_format = '0.00%'
            detail_col_offset += 5

    # ---- 日度数据表 ----
    ws_daily = wb[DAILY_DATA_SHEET]
    max_row_daily = ws_daily.max_row
    max_col_daily = ws_daily.max_column

    for col_start in range(1, max_col_daily + 1, 3):
        # 检查这3列是否有数据
        has_data = False
        for row in range(1, max_row_daily + 1):
            for col in range(col_start, min(col_start + 3, max_col_daily + 1)):
                if ws_daily.cell(row=row, column=col).value is not None:
                    has_data = True
                    break
            if has_data:
                break
        
        # 只有当这3列有数据时才合并
        if has_data:
            for row in range(1, max_row_daily + 1):
                for col in range(col_start, min(col_start + 3, max_col_daily + 1)):
                    cell_value = ws_daily.cell(row=row, column=col).value
                    merged_daily_ws.cell(row=row, column=daily_col_offset + (col - col_start), value=cell_value)
            daily_col_offset += 3

# ---- 删除列表页空行 ----
# 从最后一行向上删除空行
for row in range(merged_summary_ws.max_row, 0, -1):
    if all(merged_summary_ws.cell(row=row, column=col).value is None 
           for col in range(1, merged_summary_ws.max_column + 1)):
        merged_summary_ws.delete_rows(row)

# 保存合并后的表
merged_wb.save("合并结果.xlsx")
print("[完成] 合并完成，空行已删除，样式已保留，文件保存为：合并结果.xlsx")