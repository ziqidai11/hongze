#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# —— 在这里填写你要操作的文件路径 ——
FILE_PATHS = [
    "宏观经济/eta/1.宏观经济_数据上传.xlsx",

]




def fill_xlsx_formulas(
    file_path: str,
    summary_sheet: str = '列表页',
    detail_sheet: str = '详情页',
):
    """
    自动在 summary_sheet 中填充三列公式（直接覆盖原文件）：
      F 列 = 详情页!<pred>3
      H 列 = COUNTIF(详情页!<dir>4:<dir>30,"正确")/COUNTA(…)
      I 列 = AVERAGE(详情页!<dev>4:<dev>30)
    """
    wb = load_workbook(file_path)
    ws_sum = wb[summary_sheet]

    # 找到最后一行（以 A 列是否有值判断）
    max_row = ws_sum.max_row
    for r in range(3, max_row + 1):
        if ws_sum.cell(row=r, column=1).value is None:
            max_row = r - 1
            break

    # 从第 3 行到 max_row 逐行填公式
    for r in range(3, max_row + 1):
        base_idx = 3 + (r - 3) * 5
        pred_col = get_column_letter(base_idx)       # “预测”列
        dir_col  = get_column_letter(base_idx + 1)   # “方向”列
        dev_col  = get_column_letter(base_idx + 2)   # “偏差”列

        ws_sum.cell(row=r, column=6).value = f"=详情页!{pred_col}3"
        ws_sum.cell(row=r, column=8).value = (
            f'=COUNTIF(详情页!{dir_col}4:{dir_col}30,"正确")/'
            f'COUNTA(详情页!{dir_col}4:{dir_col}30)'
        )
        ws_sum.cell(row=r, column=9).value = f"=AVERAGE(详情页!{dev_col}4:{dev_col}30)"

    # 直接覆盖保存
    wb.save(file_path)
    print(f"✔ 已完成，已覆盖文件：{file_path}")



if __name__ == "__main__":
    for file_path in FILE_PATHS:
        fill_xlsx_formulas(file_path)

