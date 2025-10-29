import os
import calendar
from datetime import datetime, timezone, timedelta
import pandas as pd
from openpyxl import load_workbook
from copy import copy
import datetime as dt

def round_sig(x, sig=5):
    """保留有效数字"""
    try:
        if x is None:
            return None
        return float(f"{x:.{sig}g}")
    except:
        return x

def update_excel_data(
        input_data,
        excel_file_path,
        sheet_name,
        identifier,
        protected_date=None
    ):
    """
    更新Excel数据
    
    Args:
        input_data: 输入数据DataFrame
        excel_file_path: Excel文件路径
        sheet_name: 工作表名称
        identifier: 标识符
        protected_date: 保护日期（可选）
    
    Returns:
        bool: 更新是否成功
    """
    
    # 生成保护日期
    if protected_date is None:
        protected_date = _generate_protected_date()
    
    # 打开并验证工作簿
    workbook, worksheet = _open_and_validate_workbook(excel_file_path, sheet_name)
    
    # 定位标识符位置
    columns_info = _locate_identifier_columns(worksheet, identifier)
    if not columns_info:
        raise ValueError(f"未找到标识符 {identifier}")
    
    # 预处理输入数据
    processed_data = _preprocess_input_data(input_data, protected_date)
    
    # 根据当前日期决定处理策略
    current_day = datetime.now().day
    if current_day < 18:
        _handle_before_18th(worksheet, processed_data, columns_info, protected_date)
    else:
        _handle_after_18th(worksheet, processed_data, columns_info, protected_date)
    
    # 重新计算方向和偏差率
    _recalculate_metrics(worksheet, columns_info, protected_date, processed_data)
    
    # 清空表头下方的单元格
    _clear_header_cells(worksheet)
    
    # 保存文件
    workbook.save(excel_file_path)
    print(f"已完成：插入/更新数据，并重新计算\"方向/偏差率\" (protected_date = {protected_date})")
    return True

def _generate_protected_date():
    """生成保护日期"""
    jst = timezone(timedelta(hours=9))
    today = dt.datetime.now(jst)
    last_day = calendar.monthrange(today.year, today.month)[1]
    return f"{today.year:04d}/{today.month:02d}/{last_day:02d}"

def _open_and_validate_workbook(excel_file_path, sheet_name):
    """打开并验证工作簿"""
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(excel_file_path)
    
    workbook = load_workbook(excel_file_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作表 {sheet_name} 不存在")
    
    return workbook, workbook[sheet_name]

def _locate_identifier_columns(worksheet, identifier):
    """定位标识符位置并返回列信息"""
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell_value = str(worksheet.cell(row, col).value).strip()
            if cell_value == str(identifier).strip():
                return {
                    'start_row': row + 2,
                    'date_col': col,
                    'actual_col': col + 1,
                    'pred_col': col + 2,
                    'dir_col': col + 3,
                    'dev_col': col + 4
                }
    return None

def _preprocess_input_data(input_data, protected_date):
    """预处理输入数据"""
    input_data = input_data.reset_index(drop=True)
    input_data.iloc[:, 0] = input_data.iloc[:, 0].astype(str).str.strip()
    
    # 检查保护日期行
    protected_row = input_data[input_data.iloc[:, 0] == protected_date]
    if protected_row.empty:
        print(f"input_data 中缺少受保护日期行 {protected_date}")
        return None
    
    return {
        'data': input_data,
        'protected_actual': round_sig(protected_row.iloc[0].iloc[1]),
        'future_data': input_data[input_data.iloc[:, 0] > protected_date].sort_values(by=input_data.columns[0])
    }

def _write_data_to_worksheet(worksheet, data, columns_info, start_idx=0):
    """将数据写入工作表"""
    for i, row in data.iterrows():
        date_txt = str(row.iloc[0]).strip()
        actual_val = round_sig(row.iloc[1])
        pred_val = round_sig(row.iloc[2]) if len(row) > 2 else None
        
        excel_row = columns_info['start_row'] + start_idx + i
        worksheet.cell(excel_row, columns_info['date_col']).value = date_txt
        worksheet.cell(excel_row, columns_info['actual_col']).value = actual_val
        worksheet.cell(excel_row, columns_info['pred_col']).value = pred_val
        
        # 清空方向和偏差率列
        worksheet.cell(excel_row, columns_info['dir_col']).value = None
        worksheet.cell(excel_row, columns_info['dev_col']).value = None

def _handle_before_18th(worksheet, processed_data, columns_info, protected_date):
    """处理18号之前的情况"""
    if not processed_data:
        return
    
    # 只写入受保护日期及以下的数据
    filtered_data = processed_data['data'][processed_data['data'].iloc[:, 0] <= protected_date]
    _write_data_to_worksheet(worksheet, filtered_data, columns_info)

def _handle_after_18th(worksheet, processed_data, columns_info, protected_date):
    """处理18号之后的情况"""
    if not processed_data:
        return
    
    # 写入所有数据
    _write_data_to_worksheet(worksheet, processed_data['data'], columns_info)
    
    # 处理未来数据插入
    if not processed_data['future_data'].empty:
        _handle_future_data_insertion(worksheet, processed_data, columns_info, protected_date)

def _handle_future_data_insertion(worksheet, processed_data, columns_info, protected_date):
    """处理未来数据插入"""
    future_row = processed_data['future_data'].iloc[0]
    next_date = future_row.iloc[0].strip()
    next_actual = round_sig(future_row.iloc[1])
    next_pred = round_sig(future_row.iloc[2])
    
    # 找到保护日期行位置
    protected_row_idx = _find_date_row(worksheet, columns_info, protected_date)
    if not protected_row_idx:
        return
    
    # 查找下一个日期行
    next_row_idx = _find_date_row(worksheet, columns_info, next_date)
    
    if next_row_idx:
        _update_existing_next_row(worksheet, columns_info, next_row_idx, protected_row_idx, 
                                next_date, next_actual, next_pred)
    else:
        _insert_new_next_row(worksheet, columns_info, protected_row_idx, 
                           next_date, next_actual, next_pred)

def _find_date_row(worksheet, columns_info, target_date):
    """查找指定日期的行"""
    for row in range(columns_info['start_row'], worksheet.max_row + 1):
        cell_value = str(worksheet.cell(row, columns_info['date_col']).value).strip()
        if cell_value == target_date:
            return row
    return None

def _update_existing_next_row(worksheet, columns_info, next_row_idx, protected_row_idx, 
                            next_date, next_actual, next_pred):
    """更新已存在的下一个日期行"""
    worksheet.cell(next_row_idx, columns_info['actual_col']).value = next_actual
    worksheet.cell(next_row_idx, columns_info['pred_col']).value = next_pred
    
    # 如果需要移动行
    if next_row_idx > protected_row_idx - 1:
        _shift_rows_down(worksheet, columns_info, next_row_idx, protected_row_idx)
        
        # 在保护行位置写入新数据
        worksheet.cell(protected_row_idx, columns_info['date_col']).value = next_date
        worksheet.cell(protected_row_idx, columns_info['actual_col']).value = next_actual
        worksheet.cell(protected_row_idx, columns_info['pred_col']).value = next_pred
        
        # 清空原位置
        _clear_row(worksheet, columns_info, next_row_idx)

def _insert_new_next_row(worksheet, columns_info, protected_row_idx, next_date, next_actual, next_pred):
    """插入新的下一个日期行"""
    # 找到数据块的最后一行
    last_row = protected_row_idx
    while worksheet.cell(last_row, columns_info['date_col']).value not in (None, ""):
        last_row += 1
    last_row -= 1
    
    # 向下移动行
    _shift_rows_down(worksheet, columns_info, last_row, protected_row_idx)
    
    # 插入新数据
    if next_pred is not None:
        worksheet.cell(protected_row_idx, columns_info['date_col']).value = next_date
        worksheet.cell(protected_row_idx, columns_info['actual_col']).value = next_actual
        worksheet.cell(protected_row_idx, columns_info['pred_col']).value = next_pred

def _shift_rows_down(worksheet, columns_info, start_row, end_row):
    """将行向下移动"""
    for row in range(start_row, end_row - 1, -1):
        for col in [columns_info['date_col'], columns_info['actual_col'], columns_info['pred_col'],
                   columns_info['dir_col'], columns_info['dev_col']]:
            worksheet.cell(row + 1, col).value = worksheet.cell(row, col).value

def _clear_row(worksheet, columns_info, row_idx):
    """清空指定行"""
    for col in [columns_info['date_col'], columns_info['actual_col'], columns_info['pred_col'],
               columns_info['dir_col'], columns_info['dev_col']]:
        worksheet.cell(row_idx, col).value = None

def _recalculate_metrics(worksheet, columns_info, protected_date, processed_data):
    """重新计算方向和偏差率"""
    protected_row_idx = _find_date_row(worksheet, columns_info, protected_date)
    if not protected_row_idx:
        return
    
    # 更新保护行的实际值
    if processed_data and 'protected_actual' in processed_data:
        worksheet.cell(protected_row_idx, columns_info['actual_col']).value = processed_data['protected_actual']
    
    # 收集所有数据行
    data_rows = _collect_data_rows(worksheet, columns_info, protected_row_idx)
    
    # 计算偏差率和方向
    _calculate_deviation_rates(worksheet, columns_info, data_rows)
    _calculate_directions(worksheet, columns_info, data_rows)

def _collect_data_rows(worksheet, columns_info, start_row):
    """收集数据行"""
    rows = []
    current_row = start_row
    
    while True:
        date_value = worksheet.cell(current_row, columns_info['date_col']).value
        if date_value in (None, ""):
            break
            
        actual = _to_float(worksheet.cell(current_row, columns_info['actual_col']).value)
        pred = _to_float(worksheet.cell(current_row, columns_info['pred_col']).value)
        
        rows.append((current_row, actual, pred))
        current_row += 1
    
    return rows

def _to_float(value):
    """转换为浮点数"""
    try:
        return float(value) if value is not None else None
    except:
        return None

def _calculate_deviation_rates(worksheet, columns_info, data_rows):
    """计算偏差率"""
    for row_idx, actual, pred in data_rows:
        try:
            if actual is None or pred is None:
                deviation = None
            else:
                deviation = abs((pred - actual) / actual) if actual else None
            
            cell = worksheet.cell(row=row_idx, column=columns_info['dev_col'])
            cell.value = deviation
            if deviation is not None:
                cell.number_format = '0.00%'
        except Exception as e:
            print(f"计算偏差率时出错: {e}")

def _calculate_directions(worksheet, columns_info, data_rows):
    """计算方向"""
    for i, (row_idx, actual, pred) in enumerate(data_rows):
        try:
            if i == len(data_rows) - 1:
                # 最后一行不计算方向
                worksheet.cell(row_idx, columns_info['dir_col']).value = None
            else:
                _, actual_prev, pred_prev = data_rows[i + 1]
                
                if None in [actual, pred, actual_prev, pred_prev]:
                    direction = None
                else:
                    # 判断方向是否正确
                    pred_change = pred_prev - pred
                    actual_change = actual_prev - actual
                    is_correct = pred_change * actual_change >= 0
                    direction = "正确" if is_correct else "错误"
                
                worksheet.cell(row_idx, columns_info['dir_col']).value = direction
        except Exception as e:
            print(f"计算方向时出错: {e}")

def _clear_header_cells(worksheet):
    """清空表头正下方的单元格"""
    for row in range(1, worksheet.max_row):
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row, col).value
            if cell_value and isinstance(cell_value, str):
                cell_value = cell_value.strip()
                if cell_value in ["方向", "偏差率"]:
                    worksheet.cell(row + 1, col).value = None
