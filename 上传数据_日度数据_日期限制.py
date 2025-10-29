#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
FILE_PATHS = [
    "宏观经济/eta/1.宏观经济_数据上传.xlsx",
    "wti模型3.0/eta/1.WTI_数据上传.xlsx",
    "汽柴煤油2.0/eta/1.汽柴煤油_数据上传.xlsx",
    "汽柴煤油2.0/eta/2.汽柴煤油_数据上传.xlsx",
    "天然气/eta/1.天然气_数据上传.xlsx",
    "黑色/玻璃/eta/1.玻璃纯碱_数据上传.xlsx",
    "黑色/铁矿/eta/1.铁矿_数据上传.xlsx",
    "化工/eta/1.化工_数据上传.xlsx", 
    "焦煤/eta/1.焦煤_数据上传.xlsx",
    "焦炭/eta/1.焦炭_数据上传.xlsx",
    "铝/eta/1.铝_数据上传.xlsx",
    "铜/eta/1.铜_数据上传.xlsx",
    "燃料油/eta/1.燃料油_数据上传.xlsx",
    "动力煤/eta/1.动力煤_数据上传.xlsx",
    "螺纹/eta/1.螺纹_数据上传.xlsx",
    "聚丙烯(PP)/eta/1.聚丙烯_数据上传.xlsx",
    "沥青/eta/1.沥青_数据上传.xlsx"
]

# —— 可根据需要调整 —— 
SHEET_NAME = "日度数据表"
KEEP_ROWS  = 300


def trim_daily_sheets(file_paths, sheet_name=SHEET_NAME, keep_rows=KEEP_ROWS):
    """
    对列表中的每个文件：
      - 打开 workbook
      - 找到名为 sheet_name 的工作表
      - 如果行数 > keep_rows，就删除第 keep_rows+1 到 max_row 的所有行
      - 保存并覆盖原文件
    """
    for path in file_paths:
        try:
            wb = load_workbook(path)
        except FileNotFoundError:
            print(f"[警告] 文件未找到：{path}")
            continue

        if sheet_name not in wb.sheetnames:
            print(f"[警告] 工作表 \"{sheet_name}\" 不存在于 {path}")
            wb.close()
            continue

        ws = wb[sheet_name]
        max_row = ws.max_row

        if max_row > keep_rows:
            to_delete = max_row - keep_rows
            ws.delete_rows(keep_rows + 1, to_delete)
            print(f"[完成] 已在 {path} 的 \"{sheet_name}\" 删除行 {keep_rows+1}─{max_row} ({to_delete} 行)")
        else:
            print(f"[信息] {path} 的 \"{sheet_name}\" 行数 {max_row} ≤ {keep_rows}，无需删除")

        wb.save(path)
        wb.close()

    print("所有文件处理完毕")


if __name__ == "__main__":
    trim_daily_sheets(FILE_PATHS)
