import os
import calendar
from datetime import datetime, timezone, timedelta
import pandas as pd
from openpyxl import load_workbook
import datetime as dt

def filter_data_by_date(data, cutoff_date="2024/01/31"):
    """过滤数据，只保留指定日期之后的数据，并按时间倒序排列
    
    参数:
    - data: 预处理后的数据列表
    - cutoff_date: 截止日期，默认为"2024/01/31"
    
    返回:
    - 过滤并排序后的数据列表
    """
    if not data:
        return []
    
    # 将截止日期转换为datetime对象进行比较
    try:
        cutoff_dt = datetime.strptime(cutoff_date, "%Y/%m/%d")
    except ValueError:
        print(f"日期格式错误: {cutoff_date}，使用默认日期 2024/01/31")
        cutoff_dt = datetime.strptime("2024/01/31", "%Y/%m/%d")
    
    # 过滤数据
    filtered_data = []
    for item in data:
        try:
            item_date = datetime.strptime(item['date'], "%Y/%m/%d")
            if item_date >= cutoff_dt:
                filtered_data.append(item)
        except ValueError:
            print(f"跳过无效日期格式的数据: {item['date']}")
            continue
    
    # 按时间倒序排列（最新的在前）
    filtered_data.sort(key=lambda x: datetime.strptime(x['date'], "%Y/%m/%d"), reverse=True)
    
    print(f"数据过滤完成: 保留 {len(filtered_data)} 条数据（{cutoff_date} 之后），按时间倒序排列")
    return filtered_data

def round_sig(x, sig=5):
    """保留有效数字"""
    try:
        return None if x is None else float(f"{x:.{sig}g}")
    except:
        return x

def to_float(v):
    """安全转换为浮点数"""
    try:
        return float(v)
    except Exception:
        return None

def get_first_sunday_of_month(year, month):
    """获取指定月份第一个周日的日期（用于时间判断）"""
    first_day_weekday = calendar.weekday(year, month, 1)
    days_to_sunday = (6 - first_day_weekday) % 7
    first_sunday = 1 + days_to_sunday
    return f"{year:04d}/{month:02d}/{first_sunday:02d}"

def get_month_last_day(year, month):
    """获取指定月份最后一天的日期（用于数据保护）"""
    last_day = calendar.monthrange(year, month)[1]
    return f"{year:04d}/{month:02d}/{last_day:02d}"

def get_protected_date():
    """获取保护日期（当月最后一天，用于数据保护）"""
    jst = timezone(timedelta(hours=9))
    today = dt.datetime.now(jst)
    return get_month_last_day(today.year, today.month)

def should_allow_overwrite():
    """判断是否允许覆盖数据（基于当月第一个周日）"""
    jst = timezone(timedelta(hours=9))
    today = dt.datetime.now(jst)
    first_sunday = get_first_sunday_of_month(today.year, today.month)
    first_sunday_datetime = dt.datetime.strptime(first_sunday, "%Y/%m/%d")
    return today.date() < first_sunday_datetime.date()

def preprocess_data(input_data):
    """预处理输入数据
    数据结构：第一列=日期，第二列=实际值，第三列=预测值
    列名可能不固定，但位置固定
    """
    input_data = input_data.reset_index(drop=True)
    input_data.iloc[:, 0] = input_data.iloc[:, 0].astype(str).str.strip()
    
    return [
        {
            'date': str(row.iloc[0]).strip(),
            'actual': round_sig(row.iloc[1]),
            'prediction': round_sig(row.iloc[2]) if len(row) > 2 else None
        }
        for _, row in input_data.iterrows()
    ]

def find_cell_in_range(ws, value, start_row=1, end_row=None, start_col=1, end_col=None):
    """在指定范围内查找单元格"""
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column
    
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            if str(ws.cell(r, c).value).strip() == str(value).strip():
                return r, c
    return None, None

def scan_worksheet(ws, identifier, protected_date):
    """扫描工作表，一次性收集所有需要的信息"""
    # 扫描前50行找标识符
    identifier_row, identifier_col = find_cell_in_range(ws, identifier, 1, min(50, ws.max_row))
    identifier_info = (identifier_row + 2, identifier_col) if identifier_row else None
    
    # 扫描整个工作表找保护日期
    protected_row, _ = find_cell_in_range(ws, protected_date)
    
    return identifier_info, protected_row

def get_column_mapping(identifier_col):
    """获取列映射"""
    if identifier_col is None:
        return None
    
    return {
        'date': identifier_col,
        'actual': identifier_col + 1,
        'prediction': identifier_col + 2,
        'direction': identifier_col + 3,
        'deviation': identifier_col + 4
    }

def collect_data_rows(ws, start_row, columns):
    """收集数据行"""
    rows = []
    r = start_row
    
    while r <= ws.max_row:
        date_val = ws.cell(r, columns['date']).value
        if date_val in (None, ""):
            break
            
        act = to_float(ws.cell(r, columns['actual']).value)
        pred = to_float(ws.cell(r, columns['prediction']).value)
        rows.append((r, act, pred))
        r += 1
    
    return rows

def batch_write_data(ws, start_row, data, columns, max_rows=25):
    """批量写入数据"""
    for i, row_data in enumerate(data[:max_rows]):
        r = start_row + i
        ws.cell(r, columns['date']).value = row_data['date']
        ws.cell(r, columns['actual']).value = row_data['actual']
        ws.cell(r, columns['prediction']).value = row_data['prediction']
        ws.cell(r, columns['direction']).value = None
        ws.cell(r, columns['deviation']).value = None

def calculate_metrics_batch(ws, rows, columns):
    """批量计算指标"""
    for idx, (r, act, pred) in enumerate(rows):
        # 计算偏差率
        try:
            if act is None or pred is None:
                ws.cell(r, columns['deviation']).value = None
            else:
                dev_val = abs((pred - act) / act) if act else None
                cell = ws.cell(row=r, column=columns['deviation'])
                cell.value = dev_val
                cell.number_format = '0.00%'
        except Exception as e:
            print(f"计算偏差率时出错: {e}")
        
        # 计算方向
        try:
            if idx == len(rows) - 1:
                ws.cell(r, columns['direction']).value = None
            else:
                _, act_next, pred_next = rows[idx + 1]
                
                if act is None or pred is None or act_next is None or pred_next is None:
                    ws.cell(r, columns['direction']).value = None
                else:
                    ok = (act_next - pred) * (act_next - act) >= 0
                    ws.cell(r, columns['direction']).value = "正确" if ok else "错误"
        except Exception as e:
            print(f"计算方向时出错: {e}")

def clear_header_cells(ws):
    """清空表头下方的单元格"""
    for r in range(1, ws.max_row):
        for c in range(1, ws.max_column + 1):
            cell_val = ws.cell(r, c).value
            if cell_val is not None and isinstance(cell_val, str):
                cell_val = cell_val.strip()
                if cell_val in ["方向", "偏差率"]:
                    ws.cell(r + 1, c).value = None

def find_existing_row(ws, date_value, start_row, columns):
    """查找日期是否已存在"""
    for r in range(start_row, ws.max_row + 1):
        if str(ws.cell(r, columns['date']).value).strip() == date_value:
            return r
    return None

def find_first_data_row(ws, start_row, columns):
    """找到数据块开始位置"""
    first_data_row = start_row
    while ws.cell(first_data_row, columns['date']).value in (None, ""):
        first_data_row += 1
    return first_data_row

def insert_new_row(ws, data_row, start_row, columns):
    """插入新行"""
    first_data_row = find_first_data_row(ws, start_row, columns)
    
    # 从下往上移动数据，为新行腾出空间
    for r in range(ws.max_row, first_data_row - 1, -1):
        for col_name in ['date', 'actual', 'prediction', 'direction', 'deviation']:
            ws.cell(r + 1, columns[col_name]).value = ws.cell(r, columns[col_name]).value
    
    # 在顶部插入新数据
    ws.cell(first_data_row, columns['date']).value = data_row['date']
    ws.cell(first_data_row, columns['actual']).value = data_row['actual']
    ws.cell(first_data_row, columns['prediction']).value = data_row['prediction']

def process_data_row(ws, data_row, start_row, columns, allow_overwrite, protected_date):
    """处理单行数据"""
    existing_row = find_existing_row(ws, data_row['date'], start_row, columns)
    
    if existing_row:
        # 数据已存在
        # 实际值：没有保护日期，有数据就全部覆盖
        ws.cell(existing_row, columns['actual']).value = data_row['actual']
        
        # 预测值：根据保护日期策略决定是否覆盖
        if allow_overwrite and data_row['date'] >= protected_date:
            # 情况1：允许覆盖且日期在保护日期及以后
            ws.cell(existing_row, columns['prediction']).value = data_row['prediction']
        elif not allow_overwrite and data_row['date'] > protected_date:
            # 情况2：不允许覆盖但日期在保护日期之后
            ws.cell(existing_row, columns['prediction']).value = data_row['prediction']
        # 其他情况：预测值不覆盖，保持原值
    else:
        # 数据不存在，根据策略决定是否插入
        if allow_overwrite and data_row['date'] >= protected_date:
            # 情况1：允许覆盖且日期在保护日期及以后
            insert_new_row(ws, data_row, start_row, columns)
        elif not allow_overwrite and data_row['date'] > protected_date:
            # 情况2：不允许覆盖但日期在保护日期之后
            insert_new_row(ws, data_row, start_row, columns)
        # 其他情况：不插入

def update_excel_data(
        input_data,
        excel_file_path,
        sheet_name,
        identifier,
        protected_date=None,
        cutoff_date="2024/01/31"
    ):
    """更新Excel数据的主函数
    
    参数:
    - input_data: pandas DataFrame，数据结构固定：
      * 第一列：日期
      * 第二列：实际值
      * 第三列：预测值
      列名可以任意，但位置必须固定
    - excel_file_path: Excel文件路径
    - sheet_name: 工作表名称
    - identifier: 在Excel中查找的标识符（如"日期"）
    - protected_date: 保护日期（可选，默认为当月最后一天）
    - cutoff_date: 数据过滤截止日期（可选，默认为"2024/01/31"）
    
    逻辑说明：
    - 时间判断：基于当月第一个周日判断是否允许覆盖
    - 数据保护：基于当月最后一天判断哪些数据受保护
    - 数据过滤：只保留指定日期之后的数据，按时间倒序排列
    """
    
    # ---------- 0) 生成保护日期 ----------
    if protected_date is None:
        protected_date = get_protected_date()

    # ---------- 1) 打开工作簿 ----------
    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(excel_file_path)
    wb = load_workbook(excel_file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表 {sheet_name} 不存在")
    ws = wb[sheet_name]

    # ---------- 2) 扫描工作表 ----------
    identifier_info, protected_row = scan_worksheet(ws, identifier, protected_date)
    if not identifier_info:
        raise ValueError(f"未找到标识符 {identifier}")
    
    start_row, identifier_col = identifier_info
    columns = get_column_mapping(identifier_col)
    if not columns:
        raise ValueError("无法确定列映射")

    # ---------- 3) 预处理输入数据 ----------
    processed_data = preprocess_data(input_data)
    
    # ---------- 3.5) 过滤数据（只保留指定日期之后的数据，按时间倒序） ----------
    processed_data = filter_data_by_date(processed_data, cutoff_date)
    
    # 检查是否允许覆盖
    allow_overwrite = should_allow_overwrite()
    
    # 验证保护日期数据存在
    protected_data = next((d for d in processed_data if d['date'] == protected_date), None)
    if not protected_data:
        print(f"input_data 中缺少受保护日期行 {protected_date}")
        return False

    # ---------- 4) 处理数据写入 ----------
    if protected_row is None:
        # 保护日期行不存在，插入完整的月度数据
        batch_write_data(ws, start_row, processed_data, columns)
        
        # 重新查找保护日期行
        protected_row, _ = find_cell_in_range(ws, protected_date, start_row)
    else:
        # 保护日期行存在，根据覆盖策略处理
        for data_row in processed_data:
            process_data_row(ws, data_row, start_row, columns, allow_overwrite, protected_date)

    # ---------- 5) 重新计算指标 ----------
    if protected_row:
        rows = collect_data_rows(ws, protected_row, columns)
        calculate_metrics_batch(ws, rows, columns)

    # ---------- 6) 清空表头单元格 ----------
    clear_header_cells(ws)

    # ---------- 7) 保存 ----------
    wb.save(excel_file_path)
    print(f"已完成：插入/更新数据，并重新计算'方向/偏差率' (protected_date = {protected_date}, cutoff_date = {cutoff_date})")
    return True
