#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook


FILE_PATHS = [
    "宏观经济/eta/1.宏观经济_数据上传.xlsx",
    "wti模型3.0/eta/1.WTI_数据上传.xlsx",
    "汽柴煤油2.0/eta/1.汽柴煤油_数据上传.xlsx",
    "汽柴煤油2.0/eta/2.汽柴煤油_数据上传.xlsx",
    "天然气/eta/1.天然气_数据上传.xlsx",
    "黑色/玻璃/eta/1.玻璃纯碱_数据上传.xlsx",
    "黑色/铁矿/eta/1.铁矿_数据上传.xlsx",
    "化工/eta/1.化工_数据上传.xlsx", 
    "焦煤/eta/1.焦煤_数据上传.xlsx",
    "焦炭/eta/1.焦炭_数据上传.xlsx",
    "铝/eta/1.铝_数据上传.xlsx",
    "铜/eta/1.铜_数据上传.xlsx",
    "燃料油/eta/1.燃料油_数据上传.xlsx",
    "动力煤/eta/1.动力煤_数据上传.xlsx",
    "螺纹/eta/1.螺纹_数据上传.xlsx",
    "聚丙烯(PP)/eta/1.聚丙烯_数据上传.xlsx",
    "沥青/eta/1.沥青_数据上传.xlsx"
]





def fill_xlsx_values(
    file_path: str,
    summary_sheet: str = '列表页',
    detail_sheet: str = '详情页',
    start_data_row: int = 4,
    end_data_row: int = 30
):
    """
    在 summary_sheet 中填充三列计算结果，并保留数值（无公式）：
      F 列 = 详情页!<pred>3
      H 列 = COUNTIF(详情页!<dir>start_data_row:<dir>end_data_row,"正确")/
              COUNTA(详情页!<dir>start_data_row:<dir>end_data_row)
      I 列 = AVERAGE(详情页!<dev>start_data_row:<dev>end_data_row)
    """
    wb = load_workbook(file_path, data_only=True)
    ws_sum = wb[summary_sheet]
    ws_det = wb[detail_sheet]

    # 找到 summary_sheet 最后一行（以 A 列为准）
    max_row = ws_sum.max_row
    for r in range(3, max_row + 1):
        if ws_sum.cell(row=r, column=1).value is None:
            max_row = r - 1
            break

    # 遍历每一条指标（summary_sheet 第 3 行对应 detail_sheet 第 3 行）
    for r in range(3, max_row + 1):
        # detail_sheet 中第 (r-2) 个指标组的"预测"在第 3 + (r-3)*5 列
        base_idx = 3 + (r - 3) * 5
        pred_idx, dir_idx, dev_idx = base_idx, base_idx + 1, base_idx + 2

        # 读取"预测值"（固定在第 3 行）
        pred_val = ws_det.cell(row=3, column=pred_idx).value

        # 计算"方向准确率"
        dir_vals = [
            ws_det.cell(row=i, column=dir_idx).value
            for i in range(start_data_row, end_data_row + 1)
        ]
        valid_dir = [v for v in dir_vals if v not in (None, "")]
        total = len(valid_dir)
        correct = sum(1 for v in valid_dir if v == "正确")
        acc = correct / total if total else None

        # 计算"平均偏差率"
        dev_vals = [
            ws_det.cell(row=i, column=dev_idx).value
            for i in range(start_data_row, end_data_row + 1)
        ]
        numeric_dev = [v for v in dev_vals if isinstance(v, (int, float))]
        avg_dev = sum(numeric_dev) / len(numeric_dev) if numeric_dev else None

        # 写入 summary_sheet，保留数值
        cell_pred = ws_sum.cell(row=r, column=6)  # F 列
        cell_pred.value = pred_val

        cell_acc = ws_sum.cell(row=r, column=8)   # H 列
        cell_acc.value = acc
        cell_acc.number_format = '0.00%'

        cell_dev = ws_sum.cell(row=r, column=9)   # I 列
        cell_dev.value = avg_dev
        cell_dev.number_format = '0.00%'

    wb.save(file_path)
    print(f"[完成] 已处理并保存：{file_path}")


if __name__ == "__main__":
    for file_path in FILE_PATHS:
        fill_xlsx_values(file_path)
