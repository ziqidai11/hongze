
import openpyxl
from pathlib import Path

FILE_PATHS = [
    "宏观经济/eta/1.宏观经济_数据上传.xlsx",
    "wti模型3.0/eta/1.WTI_数据上传.xlsx",
    "汽柴煤油2.0/eta/1.汽柴煤油_数据上传.xlsx", 
    "天然气/eta/1.天然气_数据上传.xlsx",
    "黑色/玻璃/eta/1.玻璃纯碱_数据上传.xlsx",
    "黑色/铁矿/eta/1.铁矿_数据上传.xlsx",
    "化工/eta/1.化工_数据上传.xlsx",
    "焦煤/eta/1.焦煤_数据上传.xlsx",
    "焦炭/eta/1.焦炭_数据上传.xlsx",
    "铝/eta/1.铝_数据上传.xlsx",
    "铜/eta/1.铜_数据上传.xlsx",
    "燃料油/eta/1.燃料油_数据上传.xlsx",
    "动力煤/eta/1.动力煤_数据上传.xlsx",
    "螺纹/eta/1.螺纹_数据上传.xlsx",
    "聚丙烯(PP)/eta/1.聚丙烯_数据上传.xlsx",
    "沥青/eta/1.沥青_数据上传.xlsx"
]


def update_direction_only(
    file_path: str | Path,
    target_sheets: tuple[str, ...] = ("详情页", ),
) -> None:
    """
    只更新Excel文件中"详情页"工作表的方向列。
    方向计算方式：(act_next - pred) * (act_next - act) >= 0
    
    参数:
    - file_path: Excel 文件路径
    - target_sheets: 要处理的工作表名称元组
    """
    wb = openpyxl.load_workbook(file_path)

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"[警告] 未找到工作表: {sheet_name}（已跳过）")
            continue

        ws = wb[sheet_name]

        # 在前20行内，找到所有表头块的起始行和列
        header_blocks = []
        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value == "指标日期":
                    # 检查后续4列是否为完整表头
                    col = cell.column
                    row_idx = cell.row
                    def clean(v):
                        if v is None:
                            return ""
                        return str(v).replace(" ", "").replace("　", "")  # 去除空格和全角空格
                    values = [clean(ws.cell(row=row_idx, column=col + i).value) for i in range(5)]
                    if values == ["指标日期", "实际值", "预测值", "方向", "偏差率"]:
                        header_blocks.append((row_idx, col))
        
        if not header_blocks:
            print(f"[警告] {sheet_name} 未找到任何完整表头块，已跳过")
            continue

        def to_float(v):
            try:
                return float(v)
            except Exception:
                return None

        for header_row_idx, start_col in header_blocks:
            # 构建列名到列号的映射
            col_map = {
                "指标日期": start_col,
                "实际值": start_col + 1,
                "预测值": start_col + 2,
                "方向": start_col + 3,
                "偏差率": start_col + 4,
            }

            # 收集本块数据行
            rows = []
            r = header_row_idx + 1
            while True:
                date_txt = ws.cell(r, col_map["指标日期"]).value
                if date_txt in (None, ""):
                    break
                act = to_float(ws.cell(r, col_map["实际值"]).value)
                pred = to_float(ws.cell(r, col_map["预测值"]).value)
                rows.append((r, act, pred))
                r += 1

            # 清空旧方向值
            for r, _, _ in rows:
                ws.cell(r, col_map["方向"]).value = None

            # 重新计算方向
            for idx, (r, act, pred) in enumerate(rows):
                # 检查是否还有下一行数据（时间上更早的数据）
                if idx + 1 >= len(rows):
                    # 没有下一行数据，方向为空
                    ws.cell(r, col_map["方向"]).value = None
                else:
                    # 获取下一行数据（时间上更晚的数据作为下一行）
                    next_r, act_next, pred_next = rows[idx + 1]
                    
                    # 方向计算：(act_next - pred) * (act_next - act) >= 0
                    # 只有在当前行和下一行的数据都完整时才进行计算
                    if None in (act, pred, act_next):
                        ws.cell(r, col_map["方向"]).value = None
                    else:
                        ok = (act_next - pred) * (act_next - act) >= 0
                        ws.cell(r, col_map["方向"]).value = "正确" if ok else "错误"

    wb.save(file_path)
    print(f"[完成] 已处理并保存: {file_path}")


if __name__ == "__main__":
    for file_path in FILE_PATHS:
        update_direction_only(file_path)


