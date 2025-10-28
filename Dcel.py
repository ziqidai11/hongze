def update_excel_data(input_data, excel_file_path, sheet_name, identifier):
    """
    将数据框更新到指定Excel文件的特定工作表中
    
    参数:
        input_data (DataFrame): 要更新的数据，需包含三列（日期、实际值、预测值）
        excel_file_path (str): Excel文件路径
        sheet_name (str): 工作表名称
        identifier (str): 标识符，用于定位更新位置
        
    返回:
        bool: 更新是否成功
    """
    try:
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        import os
        from datetime import datetime
        
        # 检查文件是否存在
        if not os.path.exists(excel_file_path):
            print(f"错误：文件 {excel_file_path} 不存在")
            return False
            
        # 加载工作簿（不使用 data_only，这样我们可以访问公式）
        wb = load_workbook(excel_file_path)

        if sheet_name not in wb.sheetnames:
            print(f"错误：工作表 {sheet_name} 不存在")
            return False
            
        ws = wb[sheet_name]
        
        # 查找标识符位置
        start_row = None
        identifier_col = None
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if str(cell_value).strip() == str(identifier).strip():
                    start_row = row + 2  # 假设数据从标识符下方两行开始
                    identifier_col = col
                    break
            if start_row:
                break
                
        if not start_row:
            print(f"错误：未找到标识符 {identifier}")
            return False
            
        # 确定写入的列
        date_col = identifier_col  # 日期列在标识符下方
        actual_col = date_col + 1  # 真实值列
        pred_col = date_col + 2    # 预测值列
        
        # 只清除数据，不覆盖公式
        last_row = ws.max_row
        for row in range(start_row, last_row + 1):
            for col in [date_col, actual_col, pred_col]:
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    # 清除数据，但保留公式
                    if not isinstance(cell.value, str) or not cell.value.startswith('='):
                        cell.value = None
        
        # 写入新数据
        for i, row_data in enumerate(input_data.values):
            current_row = start_row + i
            
            # 处理日期格式
            date_value = row_data[0]
            if isinstance(date_value, (datetime, pd.Timestamp)):
                formatted_date = date_value.strftime('%Y/%m/%d')
            else:
                # 如果已经是字符串，尝试解析并重新格式化
                try:
                    formatted_date = pd.to_datetime(date_value).strftime('%Y/%m/%d')
                except:
                    formatted_date = str(date_value)  # 如果无法解析，保持原样
            
            # 更新日期、实际值、预测值
            ws.cell(row=current_row, column=date_col).value = formatted_date
            ws.cell(row=current_row, column=actual_col).value = row_data[1]
            ws.cell(row=current_row, column=pred_col).value = row_data[2]
        
        # 保存文件
        wb.save(excel_file_path)
        print(f"成功更新 {sheet_name} 中的数据")
        return True
        
    except Exception as e:
        print(f"更新数据时出错: {str(e)}")
        return False
