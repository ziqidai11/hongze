import openpyxl
from pathlib import Path


FILE_PATHS = [
    "宏观经济/eta/1.宏观经济_数据上传.xlsx",
    "wti模型3.0/eta/1.WTI_数据上传.xlsx",
    "汽柴煤油2.0/eta/1.汽柴煤油_数据上传.xlsx", 
    "天然气/eta/1.天然气_数据上传.xlsx",
    "黑色/玻璃/eta/1.玻璃纯碱_数据上传.xlsx",
    "黑色/铁矿/eta/1.铁矿_数据上传.xlsx",
    "化工/eta/1.化工_数据上传.xlsx",
    "焦煤/eta/1.焦煤_数据上传.xlsx",
    "焦炭/eta/1.焦炭_数据上传.xlsx",
    "铝/eta/1.铝_数据上传.xlsx",
    "铜/eta/1.铜_数据上传.xlsx",
    "燃料油/eta/1.燃料油_数据上传.xlsx",
    "动力煤/eta/1.动力煤_数据上传.xlsx",
    "螺纹/eta/1.螺纹_数据上传.xlsx",
    "聚丙烯(PP)/eta/1.聚丙烯_数据上传.xlsx",
    "沥青/eta/1.沥青_数据上传.xlsx"
]


def fmt_sig(value, sig: int = 4, use_scientific: bool = False):
    """
    把任意数字字符串/数值格式化为 sig 位有效数字并返回字符串。
    非数字或空值原样返回。
    
    参数:
    - value: 要格式化的值
    - sig: 有效数字位数
    - use_scientific: 是否允许科学计数法，False时强制使用标准格式
    """
    if value is None:
        return value
    try:
        # 去掉中文逗号、千分位分隔等常见符号
        num = float(str(value).replace(",", "").strip())
    except ValueError:
        return str(value) if value is not None else None  # 非数字，跳过
    
    if use_scientific:
        return format(num, f".{sig}g")  # g 格式 = 有效数字，可能使用科学计数法
    else:
        # 强制使用标准格式，不使用科学计数法
        if abs(num) >= 1e6 or (abs(num) < 1e-3 and num != 0):
            # 只有非常大或非常小的数才使用科学计数法
            return format(num, f".{sig}g")
        else:
            # 使用固定小数位数，避免科学计数法
            return format(num, f".{sig-1}f" if sig > 1 else "g")


def trim_pred_dir_dev(
    file_path: str | Path,
    target_sheets: tuple[str, ...] = ("详情页", ),
    headers: tuple[str, ...] = ("预测值", "方向", "偏差率"),
    keep_rows: dict = {"预测值": 20, "方向": 20, "偏差率": 20},
    max_total_rows = None,  # 新增：整体行数限制
    sig_figs: int = 4,  # 新增：有效数字位数
    format_headers: tuple[str, ...] = ("实际值", "预测值"),  # 新增：需要格式化的列
    use_scientific: bool = False  # 新增：是否允许科学计数法
) -> None:
    """
    对指定列从表头往下只保留指定行数的数据，其余置空。
    遇到空值时跳过不计入行数。
    同时格式化数值为指定有效数字位数。
    可选择限制整体行数。

    参数:
    - file_path: Excel 文件路径
    - target_sheets: 要处理的工作表名称元组
    - headers: 要处理的表头名称元组
    - keep_rows: 每个表头要保留的行数
    - max_total_rows: 整体行数限制，超过此行数的行将被删除
    - sig_figs: 有效数字位数
    - format_headers: 需要格式化的表头名称元组
    """
    wb = openpyxl.load_workbook(file_path)

    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"[警告] 未找到工作表: {sheet_name}（已跳过）")
            continue

        ws = wb[sheet_name]

        # 在前20行内，找到所有表头块的起始行和列
        header_blocks = []
        for row in ws.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value == "指标日期":
                    # 检查后续4列是否为完整表头
                    col = cell.column
                    row_idx = cell.row
                    def clean(v):
                        if v is None:
                            return ""
                        return str(v).replace(" ", "").replace("　", "")  # 去除空格和全角空格
                    values = [clean(ws.cell(row=row_idx, column=col + i).value) for i in range(5)]
                    if values == ["指标日期", "实际值", "预测值", "方向", "偏差率"]:
                        header_blocks.append((row_idx, col))
        if not header_blocks:
            print(f"[警告] {sheet_name} 未找到任何完整表头块，已跳过")
            continue

        for header_row_idx, start_col in header_blocks:
            # 构建列名到列号的映射
            col_map = {
                "指标日期": start_col,
                "实际值": start_col + 1,
                "预测值": start_col + 2,
                "方向": start_col + 3,
                "偏差率": start_col + 4,
            }

            # 找到数据截止行
            last_row = header_row_idx + 1
            while ws.cell(last_row, col_map["指标日期"]).value not in (None, ""):
                last_row += 1
            last_row -= 1

            # 新增：整体行数截断
            if max_total_rows is not None and last_row > header_row_idx + max_total_rows:
                # 删除超过限制的行
                rows_to_delete = last_row - (header_row_idx + max_total_rows)
                if rows_to_delete > 0:
                    ws.delete_rows(header_row_idx + max_total_rows + 1, rows_to_delete)
                last_row = header_row_idx + max_total_rows
                print(f"[信息] {sheet_name} 已截断至 {max_total_rows} 行数据")

            # 新增：格式化数值为有效数字
            for header in format_headers:
                if header in col_map:
                    col = col_map[header]
                    for r in range(header_row_idx + 1, last_row + 1):
                        cell = ws.cell(r, col)
                        cell_value = cell.value
                        if cell_value is not None:
                            cell.value = fmt_sig(cell_value, sig_figs, use_scientific)

            # 对每个需要处理的列进行截断
            for header in headers:
                if header not in col_map:
                    continue
                col = col_map[header]
                rows_to_keep = keep_rows[header]
                
                # 记录有效值的数量
                valid_count = 0
                # 从表头后第一行开始遍历
                for r in range(header_row_idx + 1, last_row + 1):
                    cell_value = ws.cell(r, col).value
                    if cell_value not in (None, ""):
                        valid_count += 1
                        # 如果超过了需要保留的行数，清空该单元格
                        if valid_count > rows_to_keep:
                            ws.cell(r, col).value = None

    wb.save(file_path)
    print(f"[完成] 已处理并保存: {file_path}")


# -------- 用法示例 ----------
if __name__ == "__main__":
    # 检查文件是否存在
    print("检查文件...")
    valid_files = []
    for file_path in FILE_PATHS:
        if Path(file_path).exists():
            valid_files.append(file_path)
            print(f"✓ {file_path}")
        else:
            print(f"✗ {file_path} (文件不存在)")
    
    if not valid_files:
        print("错误：没有找到任何有效文件！")
        exit(1)
    
    print(f"\n找到 {len(valid_files)} 个有效文件，开始处理...")
    
    # 处理文件
    for file_path in valid_files:
        try:
            print(f"\n正在处理: {file_path}")
            trim_pred_dir_dev(
                file_path,
                max_total_rows=30,  # 整体保留10行
                sig_figs=5,  # 4位有效数字
                format_headers=("实际值", "预测值"),  # 格式化实际值和预测值
                use_scientific=False  # 不使用科学计数法，避免出现7e+01这样的格式
            )
        except Exception as e:
            print(f"处理文件 {file_path} 时出错: {e}")
    
    print("\n所有文件处理完成！")

